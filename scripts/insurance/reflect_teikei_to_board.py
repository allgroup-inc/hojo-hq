"""定例KPI → ALLGRP board への反映を適正化する。
指定部門の board「実績」列を、その部門の定例【実数値】チーム合計(反映日までの累計)で置き換え、
現差異(=実績−経過必要数)・進捗(=実績/経過必要数)を実績に合わせて再計算する。
目標(予算)列・経過必要数はマスター由来のため触らない(不明な目標は推測しない)。
氏名は扱わない。出力(実データ入り)は非追跡ディレクトリへ。元ファイルは変更しない(別名保存)。"""
from __future__ import annotations
import argparse
import openpyxl
from openpyxl.utils import get_column_letter
import rollup_check as rc

YEN_DEPTS = {"ALL委託", "QCM"}  # ⑦⑧が円単位(÷10000で万円)

# board のボリューム/予算 指標行ラベル → 定例ブロック(実績の出所)
BOARD_ACTUAL_SOURCE = {
    "予算件数": "⑥",              # 件数実績=申込
    "予算ANP": "⑦",              # ANP実績=⑦(円部門は÷10000)
    "実稼働時間": "①",
    "発信数・配布数": "②",
    "キャッチ数・着座数": "③",
    "アポ・前確OK・保全数": "④",
    "訪問・後確OK数・保全完了数": "⑤",
    "申込数・開通数・戻入金額": "⑥",
}


def _teikei_reflect_day(teikei_path):
    wb = openpyxl.load_workbook(teikei_path, read_only=True, data_only=True)
    ws = wb["【実数値】A3縦"]
    x2 = ws.cell(2, 24).value  # 反映日の値
    wb.close()
    try:
        return int(str(x2).replace("日", "").strip())
    except (TypeError, ValueError):
        return None


def _find_block(ws, dept):
    header = "定例_" + dept
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == header:
                return r, c
    raise ValueError(f"部門ブロックが見つかりません: {header}")


def reflect(board_path, teikei_path, dept, out_path, reflect_day=None):
    reflect_day = reflect_day or _teikei_reflect_day(teikei_path)
    team = rc.parse_team_daily(teikei_path)
    def actual(block):
        v = rc.cumulative(team.get(block, []), reflect_day)
        if block == "⑦" and dept in YEN_DEPTS:
            v = v / 10000.0
        return v

    wb = openpyxl.load_workbook(board_path)  # 数式も保持(実際は値貼り付け)
    ws = rc.board_ws(wb)
    start_row, start_col = _find_block(ws, dept)
    actual_col = start_col + 1          # 実績列
    need_col = start_col + 2            # 経過必要数
    diff_col = start_col + 3            # 現差異
    prog_col = start_col + 4            # 進捗

    # 当該バンドの終端 = 次の「定例_」ヘッダ行の直前(隣バンドへ波及させない)
    end_row = min(start_row + 20, ws.max_row)
    for r in range(start_row + 1, min(start_row + 25, ws.max_row) + 1):
        if any(isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.strip().startswith("定例_")
               for c in range(1, ws.max_column + 1)):
            end_row = r - 1
            break

    changelog = []
    for r in range(start_row, end_row + 1):
        label = ws.cell(r, 2).value
        label = label.strip() if isinstance(label, str) else None
        if label not in BOARD_ACTUAL_SOURCE:
            continue
        block = BOARD_ACTUAL_SOURCE[label]
        # 安全弁: 定例に当該ブロックのデータが無い(全ゼロ=未使用)なら触らない(催事の⑤⑥等)
        if sum(team.get(block, [])) == 0:
            continue
        new_act = actual(block)
        old_act = ws.cell(r, actual_col).value
        need = ws.cell(r, need_col).value
        old_diff = ws.cell(r, diff_col).value
        old_prog = ws.cell(r, prog_col).value
        # 実績を反映、現差異・進捗を再計算(経過必要数=既存)
        ws.cell(r, actual_col).value = new_act
        new_diff = new_prog = None
        if isinstance(need, (int, float)):
            new_diff = new_act - need
            ws.cell(r, diff_col).value = new_diff
            if need != 0:
                new_prog = new_act / need
                ws.cell(r, prog_col).value = new_prog
        changelog.append({
            "row": r, "label": label,
            "actual_old": old_act, "actual_new": new_act,
            "diff_old": old_diff, "diff_new": new_diff,
            "prog_old": old_prog, "prog_new": new_prog,
            "need": need, "cell": f"{get_column_letter(actual_col)}{r}",
        })
    wb.save(out_path)
    return changelog, reflect_day


def render_changelog(dept, reflect_day, changelog):
    lines = [f"# {dept} 反映修正 changelog(反映日 day{reflect_day})", "",
             "実績列を委託KPI(定例チーム合計)で適正化。現差異・進捗は実績に合わせ再計算。目標(予算)列と経過必要数は未変更(マスター由来)。", "",
             "| セル | 指標 | 実績(旧→新) | 進捗(旧→新) |", "|---|---|---|---|"]
    def f(x):
        return "" if x is None else (f"{x:.2f}" if isinstance(x, float) else str(x))
    for c in changelog:
        po = f"{c['prog_old']*100:.0f}%" if isinstance(c['prog_old'], (int, float)) else "-"
        pn = f"{c['prog_new']*100:.0f}%" if isinstance(c['prog_new'], (int, float)) else "-"
        lines.append(f"| {c['cell']} | {c['label']} | {f(c['actual_old'])} → **{f(c['actual_new'])}** | {po} → {pn} |")
    return "\n".join(lines) + "\n"


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("board")
    p.add_argument("teikei")
    p.add_argument("dept")
    p.add_argument("out")
    p.add_argument("--day", type=int)
    args = p.parse_args(argv)
    cl, day = reflect(args.board, args.teikei, args.dept, args.out, args.day)
    print(render_changelog(args.dept, day, cl))
    print(f"(corrected board -> {args.out})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
