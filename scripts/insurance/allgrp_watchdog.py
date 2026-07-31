"""ALLGROUP 反映 見張り番 — 毎日の 定例KPI → 【ALLGRPデータ】反映を監視し、必要なら修正する。

ローカル実行(OneDrive同期先のPC)。Windowsタスクスケジューラで毎朝9時5分に `check` 実行を推奨。
同フォルダの rollup_check.py / audit_reflection.py / audit_all.py / reflect_teikei_to_board.py を使う。

使い方:
  python allgrp_watchdog.py check --config watchdog_config.json
  python allgrp_watchdog.py fix   --config watchdog_config.json   # バックアップ後に実績を定例KPIで修正

watchdog_config.json 例:
  {
    "kpi_dir":  "C:\\\\Users\\\\xxx\\\\OneDrive\\\\...\\\\定例_収集",
    "master":   "C:\\\\Users\\\\xxx\\\\OneDrive\\\\...\\\\定例_集計\\\\【ALLGRPデータ】.xlsx",
    "work_dir": "C:\\\\Users\\\\xxx\\\\OneDrive\\\\...\\\\監視ログ"
  }

出力(work_dir 配下):
  reports/watchdog_YYYY-MM-DD_HHMM.md   … 当日の要確認レポート
  history.csv                            … 毎回の結果を追記(異常の推移が見える)
  backups/【master】_YYYY-MM-DD_HHMM.xlsx … fix 実行時の修正前バックアップ
"""
from __future__ import annotations
import argparse, csv, glob, json, os, shutil
from datetime import datetime

import rollup_check as rc
import audit_reflection as ar
import audit_all as aa
import reflect_teikei_to_board as rt


def _now():
    return datetime.now()


def _find_master_sheet_month(master):
    import openpyxl
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    month = None
    if "【ALLGRP】A3縦" in wb.sheetnames:
        t = str(wb["【ALLGRP】A3縦"].cell(2, 2).value or "")
        for m in ("6月", "7月", "8月", "9月", "10月", "11月", "12月", "1月", "2月", "3月", "4月", "5月"):
            if m in t:
                month = m
                break
    wb.close()
    return month


def analyze(kpi_dir, master):
    """マスター(board)と定例KPIを突き合わせ、部門別の突合結果と負値/空欄異常を返す。"""
    month = _find_master_sheet_month(master)
    # 定例ファイルを中身(管轄)で同定
    teikei = {}
    for f in glob.glob(os.path.join(kpi_dir, "*.xlsx")):
        kind, m, dept = aa.classify(f)
        if kind == "teikei" and dept and (month is None or m == month):
            teikei.setdefault(dept, []).append(f)

    results = {}
    for dep in aa.ALL_DEPTS:
        paths = teikei.get(dep, [])
        if not paths:
            results[dep] = {"status": "missing", "detail": "定例KPI未検出", "path": None}
            continue
        blocks = aa.DEPT_BLOCKS.get(dep, ["①", "②", "③", "④", "⑤", "⑥"])
        try:
            ba = rc.board_dept_actuals(master, dep)
        except Exception as e:
            results[dep] = {"status": "error", "detail": f"マスター読取失敗: {e}", "path": None}
            continue
        sub = {b: ba[b] for b in blocks if b in ba}
        best = None
        for f in paths:
            team = rc.parse_team_daily(f)
            rr = rc.reconcile(team, sub)
            days = [v["matched_day"] for v in rr.values() if v["ok"]]
            modal = max(set(days), key=days.count) if days else None
            aligned = sum(1 for v in rr.values() if v["ok"] and v["matched_day"] == modal)
            if best is None or aligned > best[0]:
                best = (aligned, len(blocks), modal, f)
        aligned, need, modal, path = best
        # 段階判定: 全一致=ok / 全滅(0)=broken(反映異常・重大) / 一部=partial(日付揃えの誤差・要目視)
        if aligned == need:
            status = "ok"
        elif aligned == 0:
            status = "broken"
        else:
            status = "partial"
        results[dep] = {
            "status": status,
            "aligned": aligned, "need": need,
            "detail": f"{aligned}/{need} 一致" + (f" @day{modal}" if modal else " 一致日なし"),
            "path": path,
        }

    # 負値・空欄の一次スクリーン(部門別符号ルール込み)
    try:
        recs = ar.extract_board_records(master)
        neg = ar.detect_negative_anomalies(recs)
    except Exception:
        neg = []
    return {"month": month, "depts": results, "negatives": neg}


def render_report(stamp, analysis):
    depts = analysis["depts"]
    n_ok = sum(1 for d in depts.values() if d["status"] == "ok")
    n_broken = sum(1 for d in depts.values() if d["status"] in ("broken", "error"))
    n_partial = sum(1 for d in depts.values() if d["status"] == "partial")
    n_missing = sum(1 for d in depts.values() if d["status"] == "missing")
    lines = [f"# 反映 見張り番レポート {stamp}", "",
             f"対象月: {analysis['month'] or '不明'} / 整合 {n_ok} ・ 反映異常 {n_broken} ・ 一部要目視 {n_partial} ・ 未検出 {n_missing}", "",
             "| 部門 | 判定 | 詳細 |", "|---|---|---|"]
    mark = {"ok": "✅ 整合", "broken": "⛔ 反映異常(重大)", "partial": "🟡 一部要目視",
            "error": "⛔ エラー", "missing": "— 未検出"}
    for dep in aa.ALL_DEPTS:
        d = depts[dep]
        lines.append(f"| {dep} | {mark[d['status']]} | {d['detail']} |")
    lines.append("")
    if analysis["negatives"]:
        lines.append(f"## 負値/空欄の要確認({len(analysis['negatives'])}件)")
        lines.append("| セル | 部門 | 指標 | 値 |")
        lines.append("|---|---|---|---|")
        for x in analysis["negatives"]:
            lines.append(f"| {x['coord']} | {x.get('area','')} | {x['metric']} | {x['value']} |")
    else:
        lines.append("負値/空欄の要確認はありません。")
    return "\n".join(lines) + "\n"


def append_history(work_dir, stamp, analysis):
    path = os.path.join(work_dir, "history.csv")
    new = not os.path.exists(path)
    with open(path, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        if new:
            w.writerow(["timestamp", "month", "dept", "status", "detail"])
        for dep in aa.ALL_DEPTS:
            d = analysis["depts"][dep]
            w.writerow([stamp, analysis["month"] or "", dep, d["status"], d["detail"]])
        w.writerow([stamp, analysis["month"] or "", "_negatives", str(len(analysis["negatives"])), "負値/空欄件数"])


def do_fix(master, work_dir, analysis, stamp):
    """反映異常(broken)の部門について、バックアップを取ってから実績を定例KPIで反映修正する。
    一部要目視(partial)は日付揃えの誤差の可能性が高いため自動修正の対象にしない(誤修正回避)。"""
    fixed = []
    targets = [dep for dep, d in analysis["depts"].items() if d["status"] == "broken" and d["path"]]
    if not targets:
        return fixed, None
    backups = os.path.join(work_dir, "backups")
    os.makedirs(backups, exist_ok=True)
    backup = os.path.join(backups, f"{os.path.splitext(os.path.basename(master))[0]}_{stamp}.xlsx")
    shutil.copy2(master, backup)
    for dep in targets:
        try:
            cl, day = rt.reflect(master, analysis["depts"][dep]["path"], dep, master)
            fixed.append((dep, len(cl), day))
        except Exception as e:
            fixed.append((dep, f"失敗: {e}", None))
    return fixed, backup


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("mode", choices=["check", "fix"])
    p.add_argument("--config", required=True)
    args = p.parse_args(argv)
    cfg = json.load(open(args.config, encoding="utf-8"))
    kpi_dir, master, work_dir = cfg["kpi_dir"], cfg["master"], cfg["work_dir"]
    os.makedirs(os.path.join(work_dir, "reports"), exist_ok=True)

    stamp = _now().strftime("%Y-%m-%d_%H%M")
    analysis = analyze(kpi_dir, master)

    if args.mode == "fix":
        fixed, backup = do_fix(master, work_dir, analysis, stamp)
        if fixed:
            print(f"修正: {fixed}  (backup: {backup})")
            analysis = analyze(kpi_dir, master)  # 修正後に再評価
        else:
            print("修正対象なし(要確認の部門はありません)")

    report = render_report(stamp, analysis)
    rp = os.path.join(work_dir, "reports", f"watchdog_{stamp}.md")
    with open(rp, "w", encoding="utf-8") as f:
        f.write(report)
    append_history(work_dir, stamp, analysis)

    bad = [d for d, v in analysis["depts"].items() if v["status"] in ("broken", "error")]
    print(report)
    print(f"(report: {rp} / history 追記済)")
    # 重大異常(反映異常・負値)があれば非ゼロ終了(タスクスケジューラ等での検知用)
    return 1 if (bad or analysis["negatives"]) else 0


if __name__ == "__main__":
    raise SystemExit(main())
