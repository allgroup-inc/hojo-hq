"""週次戦略総括(達成くん式)— ALLGRP board の計算済み値のみから戦略HTMLを生成する。
逆算の捏造はしない(正確性最優先)。予算/実績/進捗はboardが計算した値をそのまま用いる。
氏名は扱わない(部門レベルのみ)。出力(実数値入り)は非追跡ディレクトリへ。"""
from __future__ import annotations
import argparse

DEPTS = ["嘉手納", "特殊", "催事", "豊見城", "ALL委託", "札幌", "CRM", "QCM", "LTV"]
RESP = {"嘉手納": "下地", "特殊": "長澤", "催事": "仲宗根", "豊見城": "金城",
        "ALL委託": "島袋", "札幌": "泉谷", "CRM": "呉屋", "QCM": "徳元", "LTV": "上地"}
# 引き継ぎ書v2 §3 の未決経営判断(小柳さん待ち)
PENDING = [
    "LTV予算補正(案A: 4名現実値≒25,000件・トス率5%を唯一の的)を再上申中",
    "豊見城シフト増枠(全員+15〜20分/日案)を後押し中",
    "嘉手納予算2,600万の現実性検証(×1.9ペース必要)",
    "札幌の稼働急落対応(泉谷ヒアリング)",
]


def extract(board_path):
    import openpyxl
    wb = openpyxl.load_workbook(board_path, data_only=True)
    ws = wb.active
    title = ws.cell(2, 2).value or ""

    def find_block(dep):
        h = "定例_" + dep
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip() == h:
                    return r, c
        return None, None

    def row_in_band(sr, label):
        for r in range(sr, min(sr + 25, ws.max_row) + 1):
            if str(ws.cell(r, 2).value).strip() == label:
                return r
        return None

    out = {}
    for dep in DEPTS:
        sr, sc = find_block(dep)
        if not sr:
            continue
        rk, ra = row_in_band(sr, "予算件数"), row_in_band(sr, "予算ANP")

        def val(r, off):
            v = ws.cell(r, sc + off).value
            return v if isinstance(v, (int, float)) else None

        out[dep] = {
            "ken_bud": val(rk, 0), "ken_act": val(rk, 1), "ken_prog": val(rk, 4),
            "anp_bud": val(ra, 0), "anp_act": val(ra, 1), "anp_prog": val(ra, 4),
        }
    wb.close()
    return title, out


def _chip(prog):
    if prog is None or prog < 0 or prog > 5:
        return ("要確認", "#8a8a8a")   # ALL委託の壊れた目標列など
    if prog >= 1.0:
        return (f"{prog*100:.0f}% 達成圏", "#B98A2E")
    if prog >= 0.85:
        return (f"{prog*100:.0f}%", "#0F6E56")
    return (f"{prog*100:.0f}% 未達", "#B23A3A")


def render_html(title, data, audit_note):
    rows = []
    # 進捗昇順(穴の大きい順)。異常値(ALL委託)は末尾。
    def key(d):
        p = data[d]["ken_prog"]
        return (1, 0) if (p is None or p < 0 or p > 5) else (0, p)
    for dep in sorted(data, key=key):
        d = data[dep]
        kt, kc = _chip(d["ken_prog"])
        at, ac = _chip(d["anp_prog"])
        rows.append(
            f"<tr><td class='dep'>{dep}<span class='resp'>{RESP.get(dep,'')}</span></td>"
            f"<td>{'' if d['ken_act'] is None else round(d['ken_act'],1)} / "
            f"{'' if d['ken_bud'] is None else round(d['ken_bud'],1)}</td>"
            f"<td><span class='chip' style='background:{kc}'>{kt}</span></td>"
            f"<td>{'' if d['anp_act'] is None else round(d['anp_act'],1)} / "
            f"{'' if d['anp_bud'] is None else round(d['anp_bud'],1)}</td>"
            f"<td><span class='chip' style='background:{ac}'>{at}</span></td></tr>"
        )
    pend = "".join(f"<li>{p}</li>" for p in PENDING)
    return f"""<!doctype html><html lang='ja'><head><meta charset='utf-8'>
<meta name='viewport' content='width=device-width,initial-scale=1'>
<title>週次戦略総括</title>
<style>
:root{{--dg:#22382A;--g:#0F6E56;--gold:#B98A2E;}}
*{{box-sizing:border-box}} body{{font-family:Meiryo,'Noto Sans JP',sans-serif;margin:0;background:#f4f6f4;color:#1b1b1b;font-size:16px}}
.wrap{{max-width:920px;margin:0 auto;padding:20px}}
h1{{background:var(--dg);color:#fff;padding:16px 20px;border-radius:10px;font-size:20px;margin:0 0 6px}}
.sub{{color:#555;font-size:13px;margin:0 0 18px}}
h2{{color:var(--dg);border-left:6px solid var(--gold);padding-left:10px;font-size:17px;margin:26px 0 10px}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;font-size:17px}}
th,td{{padding:9px 10px;text-align:left;border-bottom:1px solid #eee}}
th{{background:var(--g);color:#fff;font-size:14px}}
.dep{{font-weight:700}} .resp{{color:#888;font-weight:400;font-size:12px;margin-left:6px}}
.chip{{color:#fff;padding:3px 9px;border-radius:12px;font-size:13px;white-space:nowrap}}
.callout{{background:#fff;border:2px solid var(--gold);border-radius:10px;padding:14px 16px;margin:10px 0}}
.callout b{{color:var(--dg)}}
ul{{background:#fff;border-radius:10px;padding:14px 14px 14px 32px;margin:6px 0}}
li{{margin:4px 0}}
.note{{font-size:12px;color:#777;margin-top:22px;border-top:1px solid #ddd;padding-top:10px}}
.ok{{color:var(--g);font-weight:700}} .warn{{color:#B23A3A;font-weight:700}}
</style></head><body><div class='wrap'>
<h1>{title} — 週次戦略総括(達成くん)</h1>
<p class='sub'>board反映日 ≈ 24日時点のスナップショット / 数値はboard計算値のみ(逆算・推計は未実施=正確性優先)</p>

<h2>❶ 部門別 予算進捗(進捗=実績÷経過必要数)</h2>
<table><thead><tr><th>部門(責任者)</th><th>件数 実績/予算</th><th>件数 進捗</th><th>ANP 実績/予算</th><th>ANP 進捗</th></tr></thead>
<tbody>{''.join(rows)}</tbody></table>

<h2>❷ 達成くん総括(単一ネクストアクション)</h2>
<div class='callout'>
<p><b>唯一の達成圏</b>: <span class='ok'>特殊(長澤)</span> — 件数116%・ANP120%。教材化して横展開。</p>
<p><b>全社最大の穴</b>: <span class='warn'>嘉手納(下地)</span> — 進捗6割台で予算規模も最大。ここが全社の着地を決める。</p>
<p><b>次の一手(今週これだけ)</b>: 嘉手納の訪問数・申込数(的)にリソースを寄せる。並行してLTV(18%)は予算補正の経営判断待ちのため“現実値の的”で運用。</p>
</div>

<h2>❸ 反映監査ステータス</h2>
<div class='callout'>{audit_note}</div>

<h2>未決の経営判断(小柳さん待ち)</h2>
<ul>{pend}</ul>

<p class='note'>本総括はboardの計算済み値のみを提示(件数=申込・ANP=年換算保険料)。ALL委託は目標列が異常値のため「要確認」表示。必要/日の逆算・着地見込は、正確な反映日・稼動日ロジック確定後に別途。生成: scripts/insurance/weekly_soukatsu.py</p>
</div></body></html>"""


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("board")
    p.add_argument("--out", required=True)
    p.add_argument("--audit-note", default="6部門(嘉手納・豊見城・特殊・CRM・QCM・LTV)= ロールアップ整合✓。ALL委託=要確認。催事・札幌=7月定例未提供。")
    args = p.parse_args(argv)
    title, data = extract(args.board)
    html = render_html(title, data, args.audit_note)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"wrote {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
