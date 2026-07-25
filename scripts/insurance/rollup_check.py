"""ロールアップ照合 — 定例【実数値】の「チーム合計」= board 実績 を突合する(本命の反映監査)。
board 実績は反映日までのMTD、定例チーム合計は日次。日付を揃えて一致する日(matched_day)を探す。
純関数(cumulative/match_day/reconcile)＋xlsxアダプタ。氏名は扱わない(チーム合計行のみ使用)。"""
from __future__ import annotations
import argparse

BLOCKS = ["★", "①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧"]

BOARD_SHEET = "【ALLGRP】A3縦"


def board_ws(wb):
    """マスター(board)のシートを返す。名前で明示指定し、無ければ active。"""
    return wb[BOARD_SHEET] if BOARD_SHEET in wb.sheetnames else wb.active

# board のボリューム指標行 → 定例ブロック(①〜⑥)。部門により現場用語は違うが位置は共通。
BOARD_VOLUME_TO_BLOCK = {
    "実稼働時間": "①",
    "発信数・配布数": "②",
    "キャッチ数・着座数": "③",
    "アポ・前確OK・保全数": "④",
    "訪問・後確OK数・保全完了数": "⑤",
    "申込数・開通数・戻入金額": "⑥",
}


# ---- 純関数(TDD対象) ----

def cumulative(daily, upto_day):
    """daily(日1..N)の先頭 upto_day 日ぶんの合計。"""
    return sum(daily[:upto_day])


def match_day(daily, target, tol=0.5, rel_tol=0.0):
    """累計が target と一致する最初の日(1始まり)。無ければ None。
    許容誤差 = max(tol, rel_tol*|target|)。大きな数(配布数等)の丸め差を吸収するため相対許容も使える。"""
    lim = max(tol, rel_tol * abs(target))
    for d in range(1, len(daily) + 1):
        if abs(cumulative(daily, d) - target) <= lim:
            return d
    return None


def reconcile(team_daily, board_actuals, tol=0.5, rel_tol=0.0):
    """定例チーム合計(日次)と board 実績を突合。
    各ブロックについて board 実績が定例の何日目の累計と一致するか(matched_day)を返す。"""
    out = {}
    for block, val in board_actuals.items():
        daily = team_daily.get(block, [])
        d = match_day(daily, val, tol, rel_tol)
        out[block] = {
            "board": val,
            "matched_day": d,
            "full_total": sum(daily),
            "ok": d is not None,
        }
    return out


# ---- xlsxアダプタ(統合実行で検証) ----

def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_team_daily(path, days=31):
    """定例【実数値】A3縦から、各ブロックの「チーム合計」日次配列を取り出す。氏名行は使わない。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["【実数値】A3縦"]
    grid = [r for r in ws.iter_rows(min_row=1, max_row=300, min_col=1, max_col=40, values_only=True)]
    wb.close()
    out, cur = {}, None
    for row in grid:
        d = row[3]  # D列
        if d is None or str(d).strip() == "":
            continue
        s = str(d).strip()
        if s.startswith("Ⅱ"):
            break
        mark = next((b for b in BLOCKS if s.startswith(b)), None)
        if mark:
            cur = mark
            continue
        if cur and "チーム合計" in s:
            out[cur] = [_num(row[7 + i]) for i in range(days)]  # 7+i列 = d日
            cur = None
    return out


def board_dept_actuals(board_path, dept):
    """board から指定部門(例 'CRM')の ①〜⑥ 実績を取り出す。
    「定例_<dept>」ヘッダで部門ブロックの開始列を特定し、実績列(=開始列+1)を読む。"""
    import openpyxl
    wb = openpyxl.load_workbook(board_path, data_only=True)
    ws = board_ws(wb)
    start_col = start_row = None
    header = "定例_" + dept
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if isinstance(v, str) and v.strip() == header:
                start_col, start_row = col, row
                break
        if start_col:
            break
    if not start_col:
        wb.close()
        raise ValueError(f"部門ヘッダが見つかりません: {header}")
    actual_col = start_col + 1  # 実績列
    # 指標行はどの部門ブロックにも現れるため、当該部門ヘッダ以降の帯(row窓)に限定して読む
    out = {}
    for row in range(start_row, min(start_row + 25, ws.max_row) + 1):
        label = ws.cell(row, 2).value  # B列=指標名
        if isinstance(label, str) and label.strip() in BOARD_VOLUME_TO_BLOCK:
            block = BOARD_VOLUME_TO_BLOCK[label.strip()]
            v = ws.cell(row, actual_col).value
            if isinstance(v, (int, float)) and block not in out:
                out[block] = float(v)
    wb.close()
    return out


def render(dept, result):
    lines = [f"## {dept} ロールアップ照合", "",
             "| ブロック | board実績 | 一致した日 | 定例フル合計 | 判定 |",
             "|---|---|---|---|---|"]
    for b in ["①", "②", "③", "④", "⑤", "⑥"]:
        r = result.get(b)
        if not r:
            continue
        mark = f"day{r['matched_day']}" if r["ok"] else "**不一致**"
        lines.append(f"| {b} | {round(r['board'],1)} | {mark} | {round(r['full_total'],1)} | {'✓' if r['ok'] else '要確認'} |")
    return "\n".join(lines) + "\n"


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("board")
    p.add_argument("teikei")
    p.add_argument("dept")
    p.add_argument("--out")
    args = p.parse_args(argv)
    team = parse_team_daily(args.teikei)
    board = board_dept_actuals(args.board, args.dept)
    result = reconcile(team, board)
    report = render(args.dept, result)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"wrote {args.out}")
    else:
        print(report)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
