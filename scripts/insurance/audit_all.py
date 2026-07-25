"""統合反映監査ドライバ — アップロードされた定例/boardを自動分類・部門同定・突合する。
- 各ファイルを board / 定例 に分類し、タイトルから月を判定。
- 定例の部門は「チーム合計」がどの部門のboard実績と一致するかで機械同定(照合=同定)。
- 部門別のブロック構成を考慮(催事は①〜④、通常は①〜⑥、ALL委託は集計=直接照合対象外)。
氏名は扱わない(チーム合計行のみ)。実数値入り詳細は非追跡ディレクトリへ。"""
from __future__ import annotations
import argparse, glob, os
import rollup_check as rc

# 部門別に「board実績と突合すべきブロック」
DEPT_BLOCKS = {
    "嘉手納": ["①", "②", "③", "④", "⑤", "⑥"],
    "特殊": ["①", "②", "③", "④", "⑤", "⑥"],
    "豊見城": ["①", "②", "③", "④", "⑤", "⑥"],
    "札幌": ["①", "②", "③", "④", "⑤", "⑥"],
    "CRM": ["①", "②", "③", "④", "⑤", "⑥"],
    "QCM": ["①", "②", "③", "④", "⑤", "⑥"],
    "LTV": ["①", "②", "③", "④", "⑤", "⑥"],
    "催事": ["①", "②", "③", "④"],           # 催事は⑤⑥が無い
}
# ALL委託は「委託まとめ(集計)」で単一定例の直接ロールアップにならない → 同定/突合対象外
AGGREGATE = {"ALL委託"}
ALL_DEPTS = list(DEPT_BLOCKS) + list(AGGREGATE)


# 管轄(責任者姓)→部門。ファイル内の「【○○管轄】」から部門を確実に読むための対応表。
KANKATSU_TO_DEPT = {
    "呉屋": "CRM", "徳元": "QCM", "上地": "LTV", "島袋": "ALL委託",
    "泉谷": "札幌", "金城": "豊見城", "長澤": "特殊", "仲宗根": "催事", "下地": "嘉手納",
}


def _dept_from_kankatsu(text):
    """「【○○管轄】」等の文字列から部門名を引く。無ければ None。"""
    if not isinstance(text, str):
        return None
    for key, dep in KANKATSU_TO_DEPT.items():
        if key in text:
            return dep
    return None


def classify(path):
    """(kind, month, dept) を返す。
    kind ∈ {'board','teikei','other'}。month は '6月'/'7月'/None。
    dept は定例の場合に 【実数値】N2 / 【チーム進捗】H2 の「【○○管轄】」から判定(照合ではなく中身)。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    kind, month, dept = "other", None, None
    if "【ALLGRP】A3縦" in sheets:
        kind = "board"
        t = str(wb["【ALLGRP】A3縦"].cell(2, 2).value or "")
    elif "【実数値】A3縦" in sheets:
        kind = "teikei"
        ws = wb["【実数値】A3縦"]
        t = str(ws.cell(2, 2).value or "")
        dept = _dept_from_kankatsu(ws.cell(2, 14).value)  # N2
        if not dept and "【チーム進捗】A3縦" in sheets:
            dept = _dept_from_kankatsu(wb["【チーム進捗】A3縦"].cell(2, 8).value)  # H2
        # 月はタイトルではなく「対象月」T2(=最新対象月)で判定。タイトルは古い雛形のままの場合がある。
        t2 = ws.cell(2, 20).value  # T2
        if isinstance(t2, (int, float)):
            month = f"{int(t2)}月"
    else:
        t = ""
    if kind == "board":
        for m in ("6月", "7月"):
            if m in t:
                month = m
    wb.close()
    return kind, month, dept


def run(src, out=None):
    files = sorted(glob.glob(os.path.join(src, "*.xlsx")))
    boards = {}                       # month -> path(最後=最新想定)
    teikeis = {}                      # (month,dept) -> [paths]
    for f in files:
        kind, month, dept = classify(f)
        if kind == "board" and month:
            boards[month] = f
        elif kind == "teikei" and month and dept:
            teikeis.setdefault((month, dept), []).append(f)
    lines = ["# 統合反映監査(全月・全部門 / 部門は中身の【○○管轄】で同定)", ""]
    summary = {}
    for month, board in sorted(boards.items()):
        lines.append(f"## {month}(board: {os.path.basename(board)[:16]})")
        lines.append("| 部門 | 提出 | 突合判定 | 揃い | 一致日 |")
        lines.append("|---|---|---|---|---|")
        stat = {}
        for dep in ALL_DEPTS:
            paths = teikeis.get((month, dep), [])
            submitted = "✅" if paths else "⛔ 未提出"
            if dep in AGGREGATE:
                verdict = "⚠ 集計(直接照合対象外)" if paths else "—"
                lines.append(f"| {dep} | {submitted} | {verdict} | — | — |")
                stat[dep] = "aggregate" if paths else "missing"
                continue
            if not paths:
                lines.append(f"| {dep} | ⛔ 未提出 | — | — | — |")
                stat[dep] = "missing"
                continue
            # 提出あり → board実績と日付揃え突合(部門別ブロック)
            try:
                ba = rc.board_dept_actuals(board, dep)
            except Exception:
                ba = {}
            blocks = DEPT_BLOCKS[dep]
            sub = {b: ba[b] for b in blocks if b in ba}
            best = None
            for f in paths:
                team = rc.parse_team_daily(f)
                rr = rc.reconcile(team, sub)
                days = [v["matched_day"] for v in rr.values() if v["ok"]]
                modal = max(set(days), key=days.count) if days else None
                aligned = sum(1 for v in rr.values() if v["ok"] and v["matched_day"] == modal)
                if best is None or aligned > best[0]:
                    best = (aligned, len(blocks), modal)
            al, need, modal = best
            ok = al == need
            lines.append(f"| {dep} | {submitted} | {'✅ 整合' if ok else '⚠ 要確認'} | {al}/{need} | {('day'+str(modal)) if modal else '—'} |")
            stat[dep] = "ok" if ok else "check"
        lines.append("")
        summary[month] = stat
    report = "\n".join(lines) + "\n"
    if out:
        with open(out, "w", encoding="utf-8") as f:
            f.write(report)
    return report, summary


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("src")
    p.add_argument("--out")
    args = p.parse_args(argv)
    report, summary = run(args.src, args.out)
    print(report)
    if args.out:
        print(f"(written to {args.out})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
