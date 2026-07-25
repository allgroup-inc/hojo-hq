"""反映監査 — ALLGRP board の反映(入力)の正確性を機械検知する。
純関数(検知・照合)＋xlsxアダプタ＋CLI。断定せず「要確認」を出力する。"""
from __future__ import annotations
import argparse

# 非負であるべき指標(実数値系)。現差異・率の差など負値が正常な指標は含めない。
NONNEG_METRICS = {
    "実稼働時間", "発信数・配布数", "キャッチ数・着座数",
    "アポ・前確OK・保全数", "訪問・後確OK数・保全完了数",
    "申込数・開通数・戻入金額", "予算件数", "予算ANP",
    "時間→発信数", "想定単価",
}

# 指示書v1.0 §1/§2: QCM の解約・戻入・実効単価は「少ないほど良い」= 負値が仕様。
# ⑦⑧(ANP/戻入)は ALL委託・QCM のみ円単位、QCM⑦は絶対値。
# → 部門×指標で「負値が想定内」の組み合わせを持ち、負値スクリーンから除外する。
INVERTED_DEPTS = {"QCM"}
INVERTED_METRIC_HINTS = ("解約", "戻入", "実効単価", "想定単価")


def is_expected_negative(metric, dept):
    """指示書の符号ルール上、負値が想定内(仕様)なら True。断定を避けるための一次判定。"""
    if dept in INVERTED_DEPTS:
        m = metric or ""
        return any(h in m for h in INVERTED_METRIC_HINTS)
    return False


def detect_negative_anomalies(records):
    out = []
    for r in records:
        v = r.get("value")
        if v is None:
            continue
        if r.get("metric") not in NONNEG_METRICS or v >= 0:
            continue
        if is_expected_negative(r.get("metric"), r.get("area")):
            continue  # 仕様上の意図的マイナス(QCM等)は要確認から除外
        out.append({**r, "reason": "negative"})
    return out


def detect_blanks(records):
    return [{**r, "reason": "blank"} for r in records if r.get("value") is None]


def check_rollup(parts, total, tol=1.0):
    s = sum(p for p in parts if p is not None)
    diff = s - total
    return {"ok": abs(diff) <= tol, "sum_parts": s, "total": total, "diff": diff}


# board は C列から始まる「5列ブロック」の繰り返し。
# 各ブロック内オフセット: 0=予算 / 1=実績 / 2=経過必要数 / 3=現差異 / 4=進捗。
# 現差異・進捗(オフセット3・4)は負値が正常なので監査対象外。
_BLOCK_START = 3           # C列
_BLOCK_WIDTH = 5
_VALUE_OFFSETS = {0, 1, 2}  # 予算・実績・経過必要数のみ


def _is_value_column(col):
    if col < _BLOCK_START:
        return False
    return ((col - _BLOCK_START) % _BLOCK_WIDTH) in _VALUE_OFFSETS


def _block_index(col):
    """列→ブロック番号(0=C群,1=H群,2=M群)。"""
    return (col - _BLOCK_START) // _BLOCK_WIDTH


def extract_board_records(path):
    """ALLGRP board(xlsx)を record 形式に変換する薄いアダプタ。
    B列=指標名を拾い、非負指標の行のうち『予算/実績/経過必要数』列(現差異・進捗を除く)
    の数値セルを record 化する。各セルの部門(area)は「定例_XXX」ヘッダの直近値で解決する
    (指示書の符号ルールを部門別に適用するため)。レイアウト依存のため統合実行で実データ検証。"""
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    records = []
    metric = None
    block_dept = {}  # ブロック番号 -> 部門名(定例_XXX の直近値)
    for row in range(1, ws.max_row + 1):
        # 部門ヘッダ更新: "定例_XXX" セルを検出したら該当ブロックの部門を更新
        for col in range(_BLOCK_START, ws.max_column + 1):
            hv = ws.cell(row, col).value
            if isinstance(hv, str) and hv.strip().startswith("定例_"):
                block_dept[_block_index(col)] = hv.strip()[len("定例_"):]
        label = ws.cell(row, 2).value  # B列=指標名
        if isinstance(label, str) and label.strip():
            metric = label.strip()
        if metric not in NONNEG_METRICS:
            continue
        for col in range(_BLOCK_START, ws.max_column + 1):
            if not _is_value_column(col):
                continue
            v = ws.cell(row, col).value
            if isinstance(v, (int, float)):
                records.append({
                    "metric": metric,
                    "coord": f"{get_column_letter(col)}{row}",
                    "area": block_dept.get(_block_index(col), ""),
                    "value": float(v),
                })
    wb.close()
    return records


def render_report(anomalies):
    lines = ["# 反映監査 要確認リスト", "", "> 断定ではなく「要確認」。原文=元セルで確認してください。", ""]
    if not anomalies:
        lines.append("要確認事項は検知されませんでした。")
    else:
        lines.append("| セル | 指標 | 値 | 種別 |")
        lines.append("|---|---|---|---|")
        for x in anomalies:
            lines.append(f"| {x['coord']} | {x['metric']} | {x['value']} | {x['reason']} |")
    return "\n".join(lines) + "\n"


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("board")
    p.add_argument("--out")
    args = p.parse_args(argv)
    records = extract_board_records(args.board)
    anomalies = detect_negative_anomalies(records)
    report = render_report(anomalies)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"wrote {args.out} ({len(anomalies)} anomalies)")
    else:
        print(report)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
